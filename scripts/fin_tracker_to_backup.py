#!/usr/bin/env python3
"""Конвертер Fin_tracker.xlsx → JSON-бэкап TripPocket (формат v2).

Использование:
  python3 scripts/fin_tracker_to_backup.py "~/Downloads/Fin_tracker (1).xlsx" \
      [--app-backup trippocket-backup-YYYY-MM-DD.json] \
      [-o ~/Downloads/trippocket-import.json]

Три листа книг (Повседневные/Крупные/Квартира) становятся сферами,
лист «Доходы» — доходными транзакциями, «Справочники» — категориями.
Если передан --app-backup (JSON-экспорт из приложения), его поездки,
транзакции, категории и ручные курсы вливаются в результат — так история
из экселя объединяется с уже накопленными данными приложения.
"""

import argparse
import json
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Нужен openpyxl: pip3 install openpyxl")

HOME = "RUB"
USER = "local"

SPHERES = [
    {"id": "sphere_daily", "name": "Повседневные", "icon": "cart-outline",
     "sheet": "Повседневные", "daily_limit": 2000},
    {"id": "sphere_major", "name": "Крупные", "icon": "diamond-outline",
     "sheet": "Крупные", "daily_limit": None},
    {"id": "sphere_home", "name": "Квартира", "icon": "home-outline",
     "sheet": "Квартира", "daily_limit": None},
    # Сфера для трат «в поездке» (лист-источника нет — заполняется приложением)
    {"id": "sphere_travel", "name": "Путешествия", "icon": "airplane-outline",
     "sheet": None, "daily_limit": None},
]

INCOME_ICONS = {
    "Зарплата": "cash-outline", "Премия": "trophy-outline", "Фриланс": "laptop-outline",
    "Репетиторство": "school-outline", "Стипендия": "library-outline",
    "Кэшбэк": "card-outline", "Вклад": "trending-up-outline", "Депозит": "trending-up-outline",
    "Перевод": "swap-horizontal-outline", "Подарки": "gift-outline",
}


def slug(text: str) -> str:
    """Детерминированный id из названия («Еда вне дома» → eda-vne-doma)."""
    translit = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
        "ж": "zh", "з": "z", "и": "i", "й": "i", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    out = []
    for ch in unicodedata.normalize("NFKD", text.lower()):
        if ch in translit:
            out.append(translit[ch])
        elif ch.isalnum() and ch.isascii():
            out.append(ch)
        else:
            out.append("-")
    return re.sub(r"-+", "-", "".join(out)).strip("-")


def ts(dt: datetime) -> int:
    """Excel-дата (полночь) → unix ms в 12:00 локального дня (безопасно к TZ)."""
    return int(dt.replace(hour=12, minute=0, second=0).timestamp() * 1000)


def num(value) -> float | None:
    """Число из ячейки; терпит суммы, записанные текстом ('943.97', '1 200,50')."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(" ", "").replace(" ", "").replace(",", "."))
        except ValueError:
            return None
    return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="путь к Fin_tracker.xlsx")
    ap.add_argument("--app-backup", help="JSON-экспорт из приложения для объединения")
    ap.add_argument("-o", "--out", default="~/Downloads/trippocket-import.json")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(Path(args.xlsx).expanduser(), data_only=True)
    now = int(time.time() * 1000)

    categories: dict[str, dict] = {}  # id → category row
    transactions: list[dict] = []

    def ensure_category(name: str, kind: str, sphere_id: str | None, icon: str) -> str:
        cid = f"{kind[:3]}_{slug(name)}"
        if cid not in categories:
            categories[cid] = {
                "id": cid, "user_id": USER, "name": name, "icon": icon,
                "is_default": 0, "sort_order": len(categories),
                "kind": kind, "sphere_id": sphere_id,
            }
        return cid

    def add_tx(txid, kind, amount, cat_id, sphere_id, note, spent):
        transactions.append({
            "id": txid, "user_id": USER, "type": kind,
            "amount": round(float(amount), 2), "currency": HOME,
            "amount_home": round(float(amount), 2), "rate_home": 1,
            "amount_base": None, "rate_used": None,
            "category_id": cat_id, "sphere_id": sphere_id, "trip_id": None,
            "note": note or None, "one_time": 0,
            "spent_at": spent, "created_at": spent,
        })

    # Категории расходов из «Справочников» (чтобы сохранить порядок и пустые)
    ref = wb["Справочники"]
    ref_cols = {"Повседневные": ("B", "sphere_daily"), "Квартира": ("E", "sphere_home"),
                "Крупные": ("G", "sphere_major")}
    for col, sphere_id in ref_cols.values():
        for row in range(2, ref.max_row + 1):
            name = ref[f"{col}{row}"].value
            if isinstance(name, str) and name.strip():
                ensure_category(name.strip(), "expense", sphere_id, "pricetag-outline")

    # Повседневные: Дата | Категория | Стоимость | Комментарий
    ws = wb["Повседневные"]
    for i, (date, cat, cost, note) in enumerate(
        ws.iter_rows(min_row=2, max_col=4, values_only=True), start=2
    ):
        value = num(cost)
        if not isinstance(date, datetime) or not cat or value is None:
            continue
        cid = ensure_category(str(cat).strip(), "expense", "sphere_daily", "pricetag-outline")
        add_tx(f"xl_d_{i}", "expense", value, cid, "sphere_daily", note, ts(date))

    # Крупные и Квартира: Дата | Трата | Категория | Стоимость
    for sheet, sphere_id, prefix in (("Крупные", "sphere_major", "m"), ("Квартира", "sphere_home", "h")):
        ws = wb[sheet]
        for i, (date, name, cat, cost) in enumerate(
            ws.iter_rows(min_row=2, max_col=4, values_only=True), start=2
        ):
            value = num(cost)
            if not isinstance(date, datetime) or not cat or value is None:
                continue
            cid = ensure_category(str(cat).strip(), "expense", sphere_id, "pricetag-outline")
            add_tx(f"xl_{prefix}_{i}", "expense", value, cid, sphere_id,
                   str(name).strip() if name else None, ts(date))

    # Доходы: MMYY | Дата | Статья | Сумма | Комментарий
    ws = wb["Доходы"]
    for i, (_, date, source, amount, note) in enumerate(
        ws.iter_rows(min_row=2, max_col=5, values_only=True), start=2
    ):
        value = num(amount)
        if not isinstance(date, datetime) or not source or value is None:
            continue
        name = str(source).strip()
        cid = ensure_category(name, "income", None, INCOME_ICONS.get(name, "cash-outline"))
        add_tx(f"xl_i_{i}", "income", value, cid, None, note, ts(date))

    backup = {
        "version": 2,
        "exportedAt": now,
        "homeCurrency": HOME,
        "trips": [],
        "transactions": transactions,
        "categories": list(categories.values()),
        "spheres": [
            {"id": s["id"], "user_id": USER, "name": s["name"], "icon": s["icon"],
             "monthly_limit": None, "daily_limit": s["daily_limit"],
             "is_default": 1, "sort_order": i}
            for i, s in enumerate(SPHERES)
        ],
        "settings": [
            {"key": "language", "value": "ru"},
            {"key": "base_currency", "value": HOME},
        ],
        "rates": [],
    }

    # Объединение с экспортом приложения (поездки, их траты, ручные курсы…)
    if args.app_backup:
        app = json.loads(Path(args.app_backup).expanduser().read_text())
        backup["trips"] = app.get("trips", [])
        backup["rates"] = app.get("rates", [])
        have_tx = {t["id"] for t in transactions}
        app_txs = app.get("transactions", app.get("expenses", []))
        have_cat = {c["id"] for c in backup["categories"]}
        for c in app.get("categories", []):
            if c["id"] not in have_cat:
                c.setdefault("kind", "expense")
                c.setdefault("sphere_id", None)
                backup["categories"].append(c)
        for t in app_txs:
            if t["id"] in have_tx:
                continue
            t.setdefault("type", "expense")
            t.setdefault("amount_home", None)
            t.setdefault("rate_home", None)
            t.setdefault("sphere_id", None)
            t.setdefault("trip_id", t.pop("trip_id", None))
            t.setdefault("one_time", 0)
            backup["transactions"].append(t)
        for s in app.get("settings", []):
            if s["key"] in ("active_trip_id", "theme", "last_currency", "last_sphere"):
                backup["settings"].append(s)

    out = Path(args.out).expanduser()
    out.write_text(json.dumps(backup, ensure_ascii=False, indent=1))

    expenses = sum(1 for t in transactions if t["type"] == "expense")
    incomes = len(transactions) - expenses
    print(f"Готово: {out}")
    print(f"  расходов: {expenses}, доходов: {incomes}, категорий: {len(categories)}")
    print(f"  всего транзакций в бэкапе: {len(backup['transactions'])}")


if __name__ == "__main__":
    main()
